from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


STATUS_FILLS = {
    "Lengkap": "C6EFCE",
    "Kurang PDF": "FFC7CE",
    "Kurang Komponen": "FFEB9C",
    "Salah Folder": "F4B183",
    "Duplikat": "D9EAD3",
    "Perlu Review Manual": "D9E1F2",
    "Kode ICD Tidak Sesuai": "E6CFF2",
    "Data LIP Tidak Sesuai": "FCE4D6",
    "Berhasil": "C6EFCE",
    "Dilewati": "FFEB9C",
    "Gagal": "FFC7CE",
}


def export_review_to_excel(
    review_df: pd.DataFrame,
    orphan_pdf_df: pd.DataFrame | None,
    summary: dict[str, int],
    review_sheet_name: str = "hasil_review",
) -> bytes:
    output = BytesIO()
    summary_df = pd.DataFrame(
        [{"Metrik": key, "Nilai": value} for key, value in summary.items()]
    )

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        review_df.to_excel(writer, sheet_name=review_sheet_name[:31], index=False)
        summary_df.to_excel(writer, sheet_name="ringkasan", index=False)
        if orphan_pdf_df is not None and not orphan_pdf_df.empty:
            orphan_pdf_df.to_excel(writer, sheet_name="pdf_tanpa_excel", index=False)
        else:
            pd.DataFrame(columns=["No SEP", "Path File", "Tanggal Folder", "Sumber", "Catatan"]).to_excel(
                writer, sheet_name="pdf_tanpa_excel", index=False
            )

        workbook = writer.book
        for worksheet in workbook.worksheets:
            format_worksheet(worksheet)

    return output.getvalue()


def export_table_to_excel(
    df: pd.DataFrame,
    summary: dict[str, int] | None = None,
    *,
    sheet_name: str = "hasil",
) -> bytes:
    output = BytesIO()
    summary = summary or {}
    summary_df = pd.DataFrame(
        [{"Metrik": key, "Nilai": value} for key, value in summary.items()]
    )
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
        summary_df.to_excel(writer, sheet_name="ringkasan", index=False)
        workbook = writer.book
        for worksheet in workbook.worksheets:
            format_worksheet(worksheet)
    return output.getvalue()


def format_worksheet(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    status_col_idx = None
    for cell in worksheet[1]:
        if cell.value == "Status Akhir":
            status_col_idx = cell.column
            break

    if status_col_idx is not None:
        for row in worksheet.iter_rows(min_row=2):
            status = row[status_col_idx - 1].value
            fill_color = STATUS_FILLS.get(str(status), "")
            if fill_color:
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor=fill_color)

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)
